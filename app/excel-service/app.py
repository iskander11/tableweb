from flask import Flask, request, jsonify, send_file
from flask.json.provider import DefaultJSONProvider
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.cell_range import CellRange
import io, re, zipfile
import xml.etree.ElementTree as ET


class SafeJSONProvider(DefaultJSONProvider):
    def default(self, o):
        try:
            return super().default(o)
        except TypeError:
            return str(o)


app = Flask(__name__)
app.json_provider_class = SafeJSONProvider
app.json = SafeJSONProvider(app)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

# FortuneSheet font family index map (same as LuckySheet)
FONT_FAMILY_MAP = {
    'Times New Roman': 0,
    'Arial':           1,
    'Tahoma':          2,
    'Verdana':         3,
    'Microsoft YaHei': 4,
    '微软雅黑':          4,
    'STXihei':         5,
    'STHeiti':         6,
    'STKaiti':         7,
    'STFangsong':      8,
    'Impact':          9,
    'Courier New':     10,
    'Lucida Console':  11,
    'Consolas':        12,
    'SimSun':          13,
    '宋体':             13,
    'SimHei':          14,
    '黑体':             14,
    'Calibri':         15,
    'Cambria':         16,
    'Georgia':         17,
    'Book Antiqua':    18,
    'Palatino Linotype': 19,
}

DEFAULT_THEME_COLORS = [
    '000000', 'FFFFFF', '44546A', 'E7E6E6',
    '4472C4', 'ED7D31', 'A9D18E', 'FFC000',
    '5B9BD5', '70AD47', '0563C1', '954F72',
]

INDEXED_COLORS = {
    0: '000000', 1: 'FFFFFF', 2: 'FF0000', 3: '00FF00',
    4: '0000FF', 5: 'FFFF00', 6: 'FF00FF', 7: '00FFFF',
    8: '000000', 9: 'FFFFFF', 10: 'FF0000', 11: '00FF00',
    12: '0000FF', 13: 'FFFF00', 14: 'FF00FF', 15: '00FFFF',
    16: '800000', 17: '008000', 18: '000080', 19: '808000',
    20: '800080', 21: '008080', 22: 'C0C0C0', 23: '808080',
    24: '9999FF', 25: '993366', 26: 'FFFFCC', 27: 'CCFFFF',
    28: '660066', 29: 'FF8080', 30: '0066CC', 31: 'CCCCFF',
    32: '000080', 33: 'FF00FF', 34: 'FFFF00', 35: '00FFFF',
    36: '800080', 37: '800000', 38: '008080', 39: '0000FF',
    40: '00CCFF', 41: 'CCFFFF', 42: 'CCFFCC', 43: 'FFFF99',
    44: '99CCFF', 45: 'FF99CC', 46: 'CC99FF', 47: 'FFCC99',
    48: '3366FF', 49: '33CCCC', 50: '99CC00', 51: 'FFCC00',
    52: 'FF9900', 53: 'FF6600', 54: '666699', 55: '969696',
    56: '003366', 57: '339966', 58: '003300', 59: '333300',
    60: '993300', 61: '993366', 62: '333399', 63: '333333',
    64: None, 65: None,
}

MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
PKG_NS  = 'http://schemas.openxmlformats.org/package/2006/relationships'
OFF_NS  = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'


# ---------------------------------------------------------------------------
# Theme colors
# ---------------------------------------------------------------------------

def get_theme_colors(wb):
    try:
        if wb.loaded_theme:
            root = ET.fromstring(wb.loaded_theme)
            ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            clrScheme = root.find('.//a:clrScheme', ns)
            if clrScheme:
                colors, order = [], ['dk1','lt1','dk2','lt2','accent1','accent2',
                                     'accent3','accent4','accent5','accent6','hlink','folHlink']
                for name in order:
                    el = clrScheme.find(f'a:{name}', ns)
                    if el is not None:
                        child = list(el)
                        if child:
                            tag = child[0].tag.split('}')[-1] if '}' in child[0].tag else child[0].tag
                            if tag == 'sysClr':
                                # System color: use lastClr (actual resolved hex), not val (e.g. "WINDOW")
                                rgb = child[0].get('lastClr', '000000')
                            else:
                                rgb = child[0].get('val', '000000')
                        else:
                            rgb = '000000'
                        colors.append(rgb.upper().lstrip('#'))
                    else:
                        colors.append('000000')
                return colors
    except Exception:
        pass
    return list(DEFAULT_THEME_COLORS)


def apply_tint(hex_color, tint):
    try:
        r, g, b = int(hex_color[0:2],16), int(hex_color[2:4],16), int(hex_color[4:6],16)
        if tint >= 0:
            r,g,b = int(r+(255-r)*tint), int(g+(255-g)*tint), int(b+(255-b)*tint)
        else:
            r,g,b = int(r*(1+tint)), int(g*(1+tint)), int(b*(1+tint))
        return f'{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}'
    except Exception:
        return hex_color


def resolve_color(color_obj, theme_colors):
    if color_obj is None:
        return None
    try:
        ctype = color_obj.type
        if ctype == 'auto':
            return None  # Automatic/default color — don't override
        if ctype == 'rgb':
            argb = str(color_obj.rgb)
            # Validate: must be hex digits only (reject system color names like 'WINDOW')
            if not re.match(r'^[0-9A-Fa-f]{6,8}$', argb):
                return None
            if len(argb) == 8:
                if argb[:2] == '00':
                    return None
                return f'#{argb[2:].upper()}'
            return f'#{argb.upper()}'
        elif ctype == 'theme':
            try:
                idx = int(color_obj.theme)
                tint = float(getattr(color_obj, 'tint', 0) or 0)
                base = theme_colors[idx] if idx < len(theme_colors) else '000000'
                if tint != 0:
                    base = apply_tint(base, tint)
                return f'#{base.upper()}'
            except Exception:
                return None
        elif ctype == 'indexed':
            rgb = INDEXED_COLORS.get(color_obj.indexed)
            return f'#{rgb}' if rgb else None
    except Exception:
        pass
    return None


def resolve_fill_color(fill, theme_colors):
    if not fill:
        return None
    fill_type = getattr(fill, 'fill_type', None) or getattr(fill, 'patternType', None)
    # Only solid fills have a meaningful background color
    if fill_type != 'solid':
        return None
    fg = getattr(fill, 'fgColor', None)
    if not fg:
        return None
    # If fgColor is "auto" type, this is a default fill — skip
    if getattr(fg, 'type', None) == 'auto':
        return None
    color = resolve_color(fg, theme_colors)
    if not color:
        return None
    # Ignore white/near-white fills (Excel default "no fill" is often stored as solid white)
    if color.upper() in ('#FFFFFF', '#FEFEFE'):
        return None
    # If fgColor resolves to pure black, check bgColor:
    # Excel stores "no fill" cells as solid-black-fg + indexed:64-bg (transparent window color).
    # A genuine black fill would NOT have bgColor indexed:64.
    if color.upper() == '#000000':
        bg = getattr(fill, 'bgColor', None)
        if bg is not None:
            bg_indexed = getattr(bg, 'indexed', None)
            bg_type    = getattr(bg, 'type', None)
            if bg_type == 'indexed' and bg_indexed in (64, 65):
                return None  # Default cell state, not an explicit black fill
    return color


# ---------------------------------------------------------------------------
# Parse CF rules directly from xlsx XML (openpyxl can't read DXF colors)
# ---------------------------------------------------------------------------

def resolve_xml_color(el, theme_colors):
    """Resolve <fgColor> XML element to CSS hex string."""
    if el is None:
        return None
    rgb = el.get('rgb')
    if rgb and len(rgb) == 8 and rgb[:2] != '00':
        return f'#{rgb[2:].upper()}'
    if rgb and len(rgb) == 6:
        return f'#{rgb.upper()}'
    theme = el.get('theme')
    if theme is not None:
        idx = int(theme)
        tint = float(el.get('tint', 0) or 0)
        base = theme_colors[idx] if idx < len(theme_colors) else '000000'
        if tint != 0:
            base = apply_tint(base, tint)
        return f'#{base.upper()}'
    indexed = el.get('indexed')
    if indexed is not None:
        rgb_val = INDEXED_COLORS.get(int(indexed))
        return f'#{rgb_val}' if rgb_val else None
    return None


def _get_sheet_xml_map(content):
    """Return {sheet_name: (xml_bytes, sheet_path)} from the xlsx ZIP."""
    result = {}
    ns = MAIN_NS
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as z:
            rels_xml  = z.read('xl/_rels/workbook.xml.rels')
            rels_root = ET.fromstring(rels_xml)
            rel_map   = {r.get('Id'): r.get('Target')
                         for r in rels_root.findall(f'{{{PKG_NS}}}Relationship')}

            wb_xml  = z.read('xl/workbook.xml')
            wb_root = ET.fromstring(wb_xml)
            sheets  = wb_root.find(f'{{{ns}}}sheets')
            if sheets is None:
                return result
            for sh in sheets.findall(f'{{{ns}}}sheet'):
                name = sh.get('name', '')
                rid  = sh.get(f'{{{OFF_NS}}}id', '')
                tgt  = rel_map.get(rid, '')
                if not tgt.startswith('worksheets/'):
                    tgt = 'worksheets/' + tgt.split('worksheets/')[-1]
                try:
                    result[name] = z.read(f'xl/{tgt}')
                except Exception:
                    pass
    except Exception:
        pass
    return result


def parse_xlsx_row_heights(content):
    """Return {sheet_name: {row_0based: px_height}} from xlsx XML."""
    ns = MAIN_NS
    result = {}
    sheet_map = _get_sheet_xml_map(content)
    for sheet_name, sh_bytes in sheet_map.items():
        rh = {}
        try:
            root = ET.fromstring(sh_bytes)
            for row_el in root.findall(f'{{{ns}}}sheetData/{{{ns}}}row'):
                r   = row_el.get('r')
                ht  = row_el.get('ht')
                if r and ht:
                    try:
                        # Excel ht is in points; 1pt ≈ 1.333px at 96dpi
                        rh[int(r) - 1] = round(float(ht) * 1.333)
                    except Exception:
                        pass
        except Exception:
            pass
        result[sheet_name] = rh
    return result


def parse_xlsx_col_widths(content):
    """Return {sheet_name: {col_0based: px_width}} from xlsx XML."""
    ns = MAIN_NS
    result = {}
    sheet_map = _get_sheet_xml_map(content)
    for sheet_name, sh_bytes in sheet_map.items():
        cw = {}
        try:
            root = ET.fromstring(sh_bytes)
            cols_el = root.find(f'{{{ns}}}cols')
            if cols_el is not None:
                for col_el in cols_el.findall(f'{{{ns}}}col'):
                    min_c = int(col_el.get('min', 1))
                    max_c = int(col_el.get('max', min_c))
                    w     = col_el.get('width')
                    if w:
                        try:
                            # Excel column width unit ≈ 7px per character + 5px padding
                            px = round(float(w) * 7.5)
                            for c in range(min_c - 1, max_c):
                                cw[c] = px
                        except Exception:
                            pass
        except Exception:
            pass
        result[sheet_name] = cw
    return result


def _is_dark_color(hex_color):
    """Return True if the color is dark enough to warrant white text."""
    try:
        h = hex_color.lstrip('#')
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        luminance = 0.299*r + 0.587*g + 0.114*b
        return luminance < 130  # Dark if perceived brightness < ~51%
    except Exception:
        return False


def parse_xlsx_xf_fills(content, theme_colors):
    """
    Parse ALL fills from xl/styles.xml and map them via cellXfs to a
    per-sheet per-cell fill color, also reading table-style header colors.
    Returns: {sheet_name: {(row1based, col1based): css_color}}
    """
    ns = MAIN_NS
    result = {}

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as z:
            # --- styles.xml → fills list ---
            fills_colors = []
            xf_fills = []
            try:
                styles_xml = z.read('xl/styles.xml')
                sroot = ET.fromstring(styles_xml)

                fills_el = sroot.find(f'{{{ns}}}fills')
                if fills_el is not None:
                    for fe in fills_el.findall(f'{{{ns}}}fill'):
                        pf = fe.find(f'{{{ns}}}patternFill')
                        color = None
                        if pf is not None and pf.get('patternType') == 'solid':
                            fg = pf.find(f'{{{ns}}}fgColor')
                            color = resolve_xml_color(fg, theme_colors)
                            if not color:
                                bg = pf.find(f'{{{ns}}}bgColor')
                                color = resolve_xml_color(bg, theme_colors)
                            # Discard white/transparent/default-black fills
                            if color and color.upper() in ('#FFFFFF', '#FEFEFE'):
                                color = None
                            if color and color.upper() == '#000000':
                                bg2 = pf.find(f'{{{ns}}}bgColor')
                                if bg2 is not None:
                                    bi = getattr(bg2, 'get', lambda k, d=None: d)('indexed')
                                    if bi in ('64', '65'):
                                        color = None
                        fills_colors.append(color)

                cell_xfs = sroot.find(f'{{{ns}}}cellXfs')
                if cell_xfs is not None:
                    for xf in cell_xfs.findall(f'{{{ns}}}xf'):
                        fid = int(xf.get('fillId', 0))
                        xf_fills.append(fills_colors[fid] if fid < len(fills_colors) else None)
            except Exception:
                pass

            # --- workbook + rels → sheet file list ---
            try:
                rels_xml = z.read('xl/_rels/workbook.xml.rels')
                rels_root = ET.fromstring(rels_xml)
                rel_map = {r.get('Id'): r.get('Target')
                           for r in rels_root.findall(f'{{{PKG_NS}}}Relationship')}
            except Exception:
                rel_map = {}

            try:
                wb_xml = z.read('xl/workbook.xml')
                wb_root = ET.fromstring(wb_xml)
                sheets_el = wb_root.find(f'{{{ns}}}sheets')
                sheet_list = []
                if sheets_el is not None:
                    for sh in sheets_el.findall(f'{{{ns}}}sheet'):
                        name = sh.get('name', '')
                        rid  = sh.get(f'{{{OFF_NS}}}id', '')
                        tgt  = rel_map.get(rid, '')
                        if not tgt.startswith('worksheets/'):
                            tgt = 'worksheets/' + tgt.split('worksheets/')[-1]
                        sheet_list.append((name, tgt))
            except Exception:
                sheet_list = []

            for sheet_name, sh_path in sheet_list:
                cell_fills = {}
                try:
                    sh_xml = z.read(f'xl/{sh_path}')
                    sh_root = ET.fromstring(sh_xml)

                    # Map cell (row, col) → fill color from xf styles
                    for row_el in sh_root.findall(f'{{{ns}}}sheetData/{{{ns}}}row'):
                        ri = int(row_el.get('r', 1))
                        for c_el in row_el.findall(f'{{{ns}}}c'):
                            ref = c_el.get('r', '')
                            col_str = ''.join(ch for ch in ref if ch.isalpha())
                            try:
                                ci = column_index_from_string(col_str)
                            except Exception:
                                continue
                            s_idx = int(c_el.get('s', 0))
                            color = xf_fills[s_idx] if s_idx < len(xf_fills) else None
                            if color:
                                cell_fills[(ri, ci)] = color

                    # Also read table styles for header rows
                    sh_dir = sh_path.rsplit('/', 1)[0]
                    sh_file = sh_path.rsplit('/', 1)[-1]
                    rels_path = f'xl/{sh_dir}/_rels/{sh_file}.rels'
                    try:
                        sh_rels_xml = z.read(rels_path)
                        sh_rels = ET.fromstring(sh_rels_xml)
                        for rel in sh_rels.findall(f'{{{PKG_NS}}}Relationship'):
                            if 'table' in rel.get('Type', '').lower():
                                tbl_tgt = rel.get('Target', '')
                                if not tbl_tgt.startswith('/'):
                                    tbl_tgt = f'xl/{sh_dir}/' + tbl_tgt.lstrip('../')
                                else:
                                    tbl_tgt = tbl_tgt.lstrip('/')
                                try:
                                    tbl_xml = z.read(tbl_tgt)
                                    tbl_root = ET.fromstring(tbl_xml)
                                    ref = tbl_root.get('ref', '')
                                    header_row_cnt = int(tbl_root.get('headerRowCount', 1))
                                    style_el = tbl_root.find(f'{{{ns}}}tableStyleInfo')
                                    hdr_color = _resolve_table_header_color(
                                        tbl_root, style_el, z, ns, theme_colors)
                                    if hdr_color and ref:
                                        try:
                                            from openpyxl.worksheet.cell_range import CellRange
                                            cr = CellRange(ref)
                                            for col in range(cr.min_col, cr.max_col + 1):
                                                for row in range(cr.min_row,
                                                                 cr.min_row + header_row_cnt):
                                                    cell_fills[(row, col)] = hdr_color
                                        except Exception:
                                            pass
                                except Exception:
                                    pass
                    except Exception:
                        pass

                except Exception:
                    pass
                result[sheet_name] = cell_fills
    except Exception:
        pass
    return result


def _resolve_table_header_color(tbl_root, style_el, z, ns, theme_colors):
    """Get the header row background color from the table's style."""
    if style_el is None:
        return None
    style_name = style_el.get('name', '')
    # Built-in Office table style header colors (from most common styles)
    BUILTIN_HEADER_COLORS = {
        'TableStyleMedium1':  '#4472C4',
        'TableStyleMedium2':  '#4472C4',
        'TableStyleMedium3':  '#4472C4',
        'TableStyleMedium4':  '#4472C4',
        'TableStyleMedium5':  '#4472C4',
        'TableStyleMedium6':  '#4472C4',
        'TableStyleMedium7':  '#4472C4',
        'TableStyleMedium8':  '#70AD47',
        'TableStyleMedium9':  '#4472C4',
        'TableStyleMedium10': '#ED7D31',
        'TableStyleMedium11': '#FFC000',
        'TableStyleMedium12': '#4472C4',
        'TableStyleMedium13': '#70AD47',
        'TableStyleMedium14': '#ED7D31',
        'TableStyleMedium15': '#4472C4',
        'TableStyleMedium16': '#4472C4',
        'TableStyleMedium17': '#4472C4',
        'TableStyleMedium18': '#4472C4',
        'TableStyleMedium19': '#4472C4',
        'TableStyleMedium20': '#4472C4',
        'TableStyleMedium21': '#4472C4',
        'TableStyleDark1':    '#000000',
        'TableStyleDark2':    '#44546A',
        'TableStyleDark3':    '#44546A',
        'TableStyleLight1':   '#4472C4',
        'TableStyleLight2':   '#BFBFBF',
        'TableStyleLight3':   '#4472C4',
    }
    color = BUILTIN_HEADER_COLORS.get(style_name)
    if color:
        return color
    # Try to read custom table style from styles.xml
    try:
        styles_xml = z.read('xl/styles.xml')
        sroot = ET.fromstring(styles_xml)
        tbl_styles = sroot.find(f'{{{ns}}}tableStyles')
        if tbl_styles is not None:
            for ts in tbl_styles.findall(f'{{{ns}}}tableStyle'):
                if ts.get('name') == style_name:
                    # Find headerRow element
                    for te in ts.findall(f'{{{ns}}}tableStyleElement'):
                        if te.get('type') == 'headerRow':
                            dxf_id = te.get('dxfId')
                            if dxf_id is not None:
                                # Read DXF fill
                                dxfs = sroot.find(f'{{{ns}}}dxfs')
                                if dxfs is not None:
                                    dxf_list = dxfs.findall(f'{{{ns}}}dxf')
                                    idx = int(dxf_id)
                                    if idx < len(dxf_list):
                                        fill_el = dxf_list[idx].find(f'{{{ns}}}fill')
                                        if fill_el is not None:
                                            pf = fill_el.find(f'{{{ns}}}patternFill')
                                            if pf is not None:
                                                fg = pf.find(f'{{{ns}}}fgColor')
                                                c = resolve_xml_color(fg, theme_colors)
                                                if c:
                                                    return c
                                                bg = pf.find(f'{{{ns}}}bgColor')
                                                return resolve_xml_color(bg, theme_colors)
    except Exception:
        pass
    return None


def parse_xlsx_cf_and_styles(content, theme_colors):
    """
    Parse conditional formatting rules directly from xlsx XML.
    Returns: (dxf_colors, sheet_cf_rules)
      dxf_colors   : list of CSS color strings indexed by dxfId
      sheet_cf_rules: dict {sheet_name: [rule_dict, ...]}
    """
    ns = MAIN_NS
    dxf_colors = []
    sheet_cf_rules = {}

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as z:
            # --- styles.xml → DXF fill colors ---
            try:
                styles_xml = z.read('xl/styles.xml')
                sroot = ET.fromstring(styles_xml)
                dxfs_el = sroot.find(f'{{{ns}}}dxfs')
                if dxfs_el is not None:
                    for i, dxf_el in enumerate(dxfs_el.findall(f'{{{ns}}}dxf')):
                        # Log raw DXF XML for debugging
                        import sys
                        print(f'[DXF{i}] {ET.tostring(dxf_el, encoding="unicode")}', file=sys.stderr)
                        fill_el = dxf_el.find(f'{{{ns}}}fill')
                        color = None
                        if fill_el is not None:
                            pf = fill_el.find(f'{{{ns}}}patternFill')
                            if pf is not None:
                                fg = pf.find(f'{{{ns}}}fgColor')
                                color = resolve_xml_color(fg, theme_colors)
                                if not color:
                                    bg = pf.find(f'{{{ns}}}bgColor')
                                    color = resolve_xml_color(bg, theme_colors)
                        dxf_colors.append(color)
            except Exception as e:
                import sys
                print(f'[DXF ERROR] {e}', file=sys.stderr)

            # --- workbook relationships → sheet file paths ---
            try:
                rels_xml = z.read('xl/_rels/workbook.xml.rels')
                rels_root = ET.fromstring(rels_xml)
                rel_map = {
                    r.get('Id'): r.get('Target')
                    for r in rels_root.findall(f'{{{PKG_NS}}}Relationship')
                }
            except Exception:
                rel_map = {}

            # --- workbook.xml → sheet names and rId ---
            try:
                wb_xml = z.read('xl/workbook.xml')
                wb_root = ET.fromstring(wb_xml)
                sheets_el = wb_root.find(f'{{{ns}}}sheets')
                sheet_list = []
                if sheets_el is not None:
                    for sh in sheets_el.findall(f'{{{ns}}}sheet'):
                        name = sh.get('name', '')
                        rid  = sh.get(f'{{{OFF_NS}}}id', '')
                        target = rel_map.get(rid, '')
                        # Normalize path
                        if not target.startswith('worksheets/'):
                            target = 'worksheets/' + target.split('worksheets/')[-1]
                        sheet_list.append((name, target))
            except Exception:
                sheet_list = []

            # --- sheet XMLs → CF rules ---
            for sheet_name, target in sheet_list:
                rules = []
                try:
                    sheet_xml = z.read(f'xl/{target}')
                    sh_root = ET.fromstring(sheet_xml)
                    for cf_el in sh_root.findall(f'{{{ns}}}conditionalFormatting'):
                        sqref = cf_el.get('sqref', '')
                        for rule_el in cf_el.findall(f'{{{ns}}}cfRule'):
                            dxf_id = rule_el.get('dxfId')
                            color = None
                            if dxf_id is not None:
                                idx = int(dxf_id)
                                if idx < len(dxf_colors):
                                    color = dxf_colors[idx]
                            formulas = [f.text or '' for f in rule_el.findall(f'{{{ns}}}formula')]
                            rules.append({
                                'sqref':    sqref,
                                'type':     rule_el.get('type', ''),
                                'operator': rule_el.get('operator', ''),
                                'text':     rule_el.get('text', ''),
                                'formulas': formulas,
                                'color':    color,
                                'priority': int(rule_el.get('priority', 999)),
                            })
                except Exception:
                    pass
                sheet_cf_rules[sheet_name] = sorted(rules, key=lambda r: r['priority'])

    except Exception:
        pass

    return dxf_colors, sheet_cf_rules


def matches_cf_rule(rule, cell_value, ri, ci):
    """Return True if this CF rule applies to the cell."""
    rtype    = rule['type']
    operator = rule['operator']
    text     = rule.get('text', '') or ''
    formulas = rule.get('formulas', [])
    val_str  = str(cell_value) if cell_value is not None else ''

    try:
        if rtype in ('containsText', 'containstext'):
            return text.lower() in val_str.lower()
        if rtype in ('notContainsText', 'notcontainstext'):
            return text.lower() not in val_str.lower()
        if rtype in ('beginsWith', 'beginswith'):
            return val_str.lower().startswith(text.lower())
        if rtype in ('endsWith', 'endswith'):
            return val_str.lower().endswith(text.lower())
        if rtype in ('containsBlanks', 'containsblanks'):
            return cell_value is None or val_str.strip() == ''
        if rtype in ('notContainsBlanks', 'notcontainsblanks'):
            return cell_value is not None and val_str.strip() != ''
        if rtype == 'cellIs':
            thresh = formulas[0] if formulas else ''
            try:
                vn, tn = float(cell_value), float(thresh)
                if operator == 'greaterThan':          return vn > tn
                if operator == 'lessThan':             return vn < tn
                if operator == 'greaterThanOrEqual':   return vn >= tn
                if operator == 'lessThanOrEqual':      return vn <= tn
                if operator == 'equal':                return vn == tn
                if operator == 'notEqual':             return vn != tn
                if operator == 'between' and len(formulas) >= 2:
                    return float(formulas[0]) <= vn <= float(formulas[1])
            except (ValueError, TypeError):
                ts = str(thresh).strip('"')
                if operator == 'equal':    return val_str == ts
                if operator == 'notEqual': return val_str != ts
        if rtype == 'expression':
            formula = formulas[0] if formulas else ''
            return eval_cf_formula_xml(formula, val_str, cell_value)
    except Exception:
        pass
    return False


def eval_cf_formula_xml(formula, val_str, cell_value):
    """Evaluate simple CF formula against cell value."""
    f = formula.strip()
    # NOT(ISERROR(SEARCH("text", ref))) → contains text
    m = re.search(r'SEARCH\s*\(\s*"([^"]*)"', f, re.IGNORECASE)
    if m:
        return m.group(1).lower() in val_str.lower()
    # =ref="text"
    m = re.match(r'=?\$?[A-Z]+\$?\d+\s*=\s*"([^"]*)"', f, re.IGNORECASE)
    if m:
        return val_str == m.group(1)
    # =ref<>"text"
    m = re.match(r'=?\$?[A-Z]+\$?\d+\s*<>\s*"([^"]*)"', f, re.IGNORECASE)
    if m:
        return val_str != m.group(1)
    # =ref>value
    m = re.match(r'=?\$?[A-Z]+\$?\d+\s*(>=|<=|>|<|=|<>)\s*(-?[\d.]+)', f)
    if m:
        op, thresh = m.group(1), float(m.group(2))
        try:
            vn = float(cell_value)
            return {'=': vn==thresh,'<>': vn!=thresh,'>': vn>thresh,
                    '<': vn<thresh,'>=': vn>=thresh,'<=': vn<=thresh}.get(op, False)
        except Exception:
            pass
    return False


def get_cf_color(ri, ci, cell_value, sheet_name, sheet_cf_rules):
    """Return CF fill color for cell (ri, ci) or None."""
    rules = sheet_cf_rules.get(sheet_name, [])
    for rule in rules:
        if not rule['color']:
            continue
        sqref = rule['sqref']
        in_range = False
        for part in sqref.split():
            try:
                cr = CellRange(part)
                if cr.min_row <= ri <= cr.max_row and cr.min_col <= ci <= cr.max_col:
                    in_range = True
                    break
            except Exception:
                pass
        if in_range and matches_cf_rule(rule, cell_value, ri, ci):
            return rule['color']
    return None


# ---------------------------------------------------------------------------
# Cell value extraction
# ---------------------------------------------------------------------------

def cell_value_and_type(cell):
    from datetime import datetime, date, time, timedelta
    from openpyxl.utils.datetime import to_excel
    val = cell.value
    if val is None:
        return None, None, None, ''
    if isinstance(val, bool):
        return val, None, 'b', 'TRUE' if val else 'FALSE'
    if isinstance(val, str) and val.startswith('='):
        return None, val[1:], None, ''
    if isinstance(val, time):
        s = val.hour*3600 + val.minute*60 + val.second
        return s/86400.0, None, 'd', val.strftime('%H:%M:%S')
    if isinstance(val, timedelta):
        return val.total_seconds()/86400.0, None, 'n', str(val)
    if isinstance(val, (datetime, date)):
        try:
            serial = to_excel(val)
            fmt = (cell.number_format or '').split(';')[0]
            display = val.strftime('%d.%m.%y %H:%M') if isinstance(val,datetime) and 'h' in fmt.lower() else (val.date() if isinstance(val,datetime) else val).strftime('%d.%m.%y')
            return serial, None, 'd', display
        except Exception:
            return str(val), None, 's', str(val)
    if isinstance(val, (int, float)):
        return val, None, 'n', str(val)
    display = str(val)
    return display, None, 's', display


def get_border_style(side, theme_colors):
    if not side or not side.border_style:
        return None
    bs = side.border_style
    if bs in ('thick',):           style = 3
    elif bs in ('medium', 'mediumDashed', 'mediumDashDot', 'mediumDashDotDot'): style = 2
    elif bs in ('double',):        style = 6
    elif bs in ('dashed',):        style = 4
    elif bs in ('dotted',):        style = 5
    elif bs in ('dashDot',):       style = 7
    elif bs in ('dashDotDot',):    style = 8
    else:                          style = 1  # thin
    color = resolve_color(getattr(side,'color',None), theme_colors) or '#000000'
    # Don't apply white/light border colors (they'd be invisible)
    if not _is_dark_color(color) and color.upper() not in ('#000000','#333333','#444444','#555555','#666666'):
        color = '#000000'
    return {'style': style, 'color': color}


# ---------------------------------------------------------------------------
# Parse endpoint
# ---------------------------------------------------------------------------

@app.route('/parse', methods=['POST'])
def parse_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    content = request.files['file'].read()

    try:
        wb   = load_workbook(io.BytesIO(content), data_only=True,  rich_text=False)
        wb_f = load_workbook(io.BytesIO(content), data_only=False, rich_text=False)
    except Exception as e:
        return jsonify({'error': f'Failed to load workbook: {str(e)}'}), 400

    theme_colors = get_theme_colors(wb)
    import sys

    # Parse CF rules and all cell fills directly from xlsx XML
    dxf_colors, sheet_cf_rules = parse_xlsx_cf_and_styles(content, theme_colors)
    sheet_xml_fills = parse_xlsx_xf_fills(content, theme_colors)
    xml_row_heights = parse_xlsx_row_heights(content)
    xml_col_widths  = parse_xlsx_col_widths(content)

    sheets = []

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        ws_f = wb_f[sheet_name]

        # Skip chart sheets
        try:
            _ = ws.column_dimensions
        except AttributeError:
            continue

        column_widths, row_heights, merges = {}, {}, {}
        slave_cells = set()
        slave_master = {}  # (row1based, col1based) -> (master_row0based, master_col0based)

        # Use XML-parsed column widths (more accurate than openpyxl)
        column_widths.update(xml_col_widths.get(sheet_name, {}))

        # Use XML-parsed row heights (openpyxl misses auto-height rows)
        row_heights.update(xml_row_heights.get(sheet_name, {}))

        for merge_range in ws.merged_cells.ranges:
            r  = merge_range.min_row - 1
            c  = merge_range.min_col - 1
            rs = merge_range.max_row - merge_range.min_row + 1
            cs = merge_range.max_col - merge_range.min_col + 1
            if rs > 1 or cs > 1:
                merges[f'{r}_{c}'] = {'r': r, 'c': c, 'rs': rs, 'cs': cs}
            for ri in range(merge_range.min_row, merge_range.max_row + 1):
                for ci in range(merge_range.min_col, merge_range.max_col + 1):
                    if ri != merge_range.min_row or ci != merge_range.min_col:
                        slave_cells.add((ri, ci))
                        slave_master[(ri, ci)] = (r, c)

        filter_config = None
        if ws.auto_filter and ws.auto_filter.ref:
            filter_config = {'ref': ws.auto_filter.ref}

        cells = {}
        borders = []  # FortuneSheet config.borderInfo entries

        for row in ws.iter_rows():
            for cell in row:
                ri, ci = cell.row, cell.column

                # Borders -> config.borderInfo (FortuneSheet ignores per-cell 'bd').
                # Done for every cell, incl. merge slaves, so merged-region edges
                # (right/bottom borders live on the slave cells in Excel) survive.
                bcell = cell.border
                if bcell:
                    bt = get_border_style(bcell.top,    theme_colors)
                    bb = get_border_style(bcell.bottom, theme_colors)
                    bl = get_border_style(bcell.left,   theme_colors)
                    br = get_border_style(bcell.right,  theme_colors)
                    if bt or bb or bl or br:
                        bval = {'row_index': ri - 1, 'col_index': ci - 1}
                        if bt: bval['t'] = bt
                        if bb: bval['b'] = bb
                        if bl: bval['l'] = bl
                        if br: bval['r'] = br
                        borders.append({'rangeType': 'cell', 'value': bval})

                if (ri, ci) in slave_cells:
                    # FortuneSheet needs merged-area slave cells pointing at the master
                    mr, mc = slave_master[(ri, ci)]
                    cells[f'{ri - 1}_{ci - 1}'] = {'mc': {'r': mr, 'c': mc}}
                    continue
                r, c = ri - 1, ci - 1

                formula_cell = ws_f.cell(row=ri, column=ci)
                formula_raw  = formula_cell.value
                has_formula  = isinstance(formula_raw, str) and formula_raw.startswith('=')

                val, formula, cell_type, display = cell_value_and_type(cell)
                if has_formula:
                    formula = formula_raw[1:]

                # CF color (XML-parsed, correct) overrides base fill
                bg_color = get_cf_color(ri, ci, cell.value, sheet_name, sheet_cf_rules)
                if not bg_color:
                    # Use XML-parsed fills (includes table header colors openpyxl misses)
                    xml_fills = sheet_xml_fills.get(sheet_name, {})
                    bg_color = xml_fills.get((ri, ci))
                if not bg_color:
                    bg_color = resolve_fill_color(cell.fill, theme_colors)

                # Skip completely empty & unstyled cells
                has_style = bool(bg_color)
                if not has_style and cell.font:
                    f = cell.font
                    if f.bold or f.italic or f.underline or f.color:
                        has_style = True
                if not has_style and cell.border:
                    b = cell.border
                    if any([b.top and b.top.border_style, b.bottom and b.bottom.border_style,
                            b.left and b.left.border_style, b.right and b.right.border_style]):
                        has_style = True
                if not has_style and cell.alignment:
                    a = cell.alignment
                    if a.horizontal or a.vertical or a.wrap_text:
                        has_style = True

                is_master = f'{r}_{c}' in merges
                if val is None and not formula and not has_style and not is_master:
                    continue

                cd = {}
                if is_master:
                    m = merges[f'{r}_{c}']
                    cd['mc'] = {'r': r, 'c': c, 'rs': m['rs'], 'cs': m['cs']}
                if val is not None: cd['v'] = val
                if display:         cd['m'] = display
                if formula:         cd['f'] = formula
                if cell_type == 'd':
                    fmt = (cell.number_format or '').split(';')[0]
                    cd['ct'] = {'fa': fmt, 't': 'd'}
                    cd['t']  = 'n'
                elif cell_type:
                    cd['t'] = cell_type

                if cell_type != 'd' and cell.number_format and cell.number_format not in ('General','@','general'):
                    cd['ct'] = {'fa': cell.number_format.split(';')[0]}

                if bg_color:
                    cd['bg'] = bg_color

                font = cell.font
                if font:
                    if font.bold:      cd['bl'] = 1
                    if font.italic:    cd['it'] = 1
                    if font.underline: cd['un'] = 1
                    if font.size:      cd['fs'] = font.size
                    if font.name:
                        # Always emit the raw family name as a string. FortuneSheet
                        # renders a string ff directly via CSS font-family, so any
                        # font (incl. custom uploaded ones) works once its @font-face
                        # / FontFace is registered on the client.
                        cd['ff'] = font.name
                    if font.color:
                        fc = resolve_color(font.color, theme_colors)
                        if fc:
                            is_light_fc = fc.upper() in ('#FFFFFF','#FEFEFE','#FDFDFD','#FCFCFC','#F0F0F0')
                            if not is_light_fc:
                                cd['fc'] = fc
                            elif bg_color and _is_dark_color(bg_color):
                                # White text only on genuinely dark backgrounds
                                cd['fc'] = fc

                align = cell.alignment
                if align:
                    ht_map = {'left':1,'center':0,'right':2,'general':1,'justify':1,'distributed':1,'fill':1}
                    vt_map = {'center':0,'top':1,'bottom':2,'justify':0,'distributed':0}
                    if align.horizontal and align.horizontal in ht_map:
                        cd['ht'] = ht_map[align.horizontal]
                    if align.vertical and align.vertical in vt_map:
                        cd['vt'] = vt_map[align.vertical]
                    # FortuneSheet compares tb as a STRING ("2"); a number never matches
                    if align.wrap_text:
                        cd['tb'] = '2'
                    # Text rotation -> FortuneSheet tr (mode) / rt (arbitrary angle).
                    # Excel: 0=horizontal, 1..90=counter-clockwise (up),
                    #        91..180=clockwise (down, angle = value-90), 255=vertical stacked.
                    tr_val = getattr(align, 'textRotation', None)
                    if tr_val is None:
                        tr_val = getattr(align, 'text_rotation', None)
                    if tr_val:
                        if tr_val == 255:
                            cd['tr'] = '3'            # vertical stacked text
                        elif tr_val == 90:
                            cd['tr'] = '4'            # rotate up 90°
                        elif tr_val == 45:
                            cd['tr'] = '1'            # tilt up 45°
                        elif 0 < tr_val < 90:
                            cd['rt'] = tr_val         # arbitrary up angle
                        elif 90 < tr_val <= 180:
                            cd['rt'] = 90 - tr_val    # arbitrary down angle (negative)

                # (borders are collected into `borders` at the top of the loop)

                cells[f'{r}_{c}'] = cd

        sheet_obj = {
            'name':         sheet_name,
            'cells':        cells,
            'columnWidths': column_widths,
            'rowHeights':   row_heights,
            'merges':       merges,
            'borderInfo':   borders,
        }
        if filter_config:
            sheet_obj['filter'] = filter_config
        sheets.append(sheet_obj)

    return jsonify({'sheets': sheets})


# ---------------------------------------------------------------------------
# Export endpoint
# ---------------------------------------------------------------------------

@app.route('/export', methods=['POST'])
def export_excel():
    data = request.get_json()
    sheets_data = data.get('sheets', [])
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in sheets_data:
        ws = wb.create_sheet(title=sheet.get('name', 'Sheet1'))

        for col_idx_str, px in (sheet.get('columnWidths') or {}).items():
            ws.column_dimensions[get_column_letter(int(col_idx_str)+1)].width = round(px/7.5, 1)

        for row_idx_str, px in (sheet.get('rowHeights') or {}).items():
            ws.row_dimensions[int(row_idx_str)+1].height = round(px/1.333, 1)

        for key, cd in (sheet.get('cells') or {}).items():
            r_str, c_str = key.split('_')
            cell = ws.cell(row=int(r_str)+1, column=int(c_str)+1)

            if cd.get('f'):
                cell.value = f"={cd['f']}"
            elif cd.get('ct',{}).get('t') == 'd' and cd.get('v') is not None:
                from openpyxl.utils.datetime import from_excel
                try:
                    cell.value = from_excel(cd['v'])
                    cell.number_format = cd.get('ct',{}).get('fa','DD.MM.YY')
                except Exception:
                    cell.value = cd.get('m','')
            elif cd.get('v') is not None:
                cell.value = cd['v']

            if cd.get('bg'):
                cell.fill = PatternFill(fill_type='solid', fgColor='FF'+cd['bg'].lstrip('#').upper())

            fkw = {}
            if cd.get('bl'): fkw['bold']      = True
            if cd.get('it'): fkw['italic']    = True
            if cd.get('un'): fkw['underline'] = 'single'
            if cd.get('fs'): fkw['size']      = cd['fs']
            if isinstance(cd.get('ff'), str) and cd['ff'].strip():
                fkw['name'] = cd['ff']
            if cd.get('fc'): fkw['color']     = 'FF'+cd['fc'].lstrip('#').upper()
            if fkw: cell.font = Font(**fkw)

            akw = {}
            ht_rev = {0:'center',1:'left',2:'right'}
            vt_rev = {0:'center',1:'top',2:'bottom'}

            def _as_int(x):
                try:    return int(x)
                except (TypeError, ValueError): return None

            ht_i = _as_int(cd.get('ht'))
            vt_i = _as_int(cd.get('vt'))
            if ht_i is not None: akw['horizontal'] = ht_rev.get(ht_i, 'left')
            if vt_i is not None: akw['vertical']   = vt_rev.get(vt_i, 'center')
            if str(cd.get('tb')) == '2': akw['wrap_text'] = True

            # Rotation: rt (arbitrary angle) takes priority over tr (discrete mode)
            rt_i = _as_int(cd.get('rt'))
            if rt_i is not None and rt_i != 0:
                akw['textRotation'] = rt_i if rt_i > 0 else (90 - rt_i)
            elif cd.get('tr') is not None:
                tr_map = {'3': 255, '4': 90, '1': 45, '2': 135, '5': 180}
                tr_excel = tr_map.get(str(cd['tr']))
                if tr_excel is not None:
                    akw['textRotation'] = tr_excel

            if akw: cell.alignment = Alignment(**akw)

        # Borders come from config.borderInfo (FortuneSheet format)
        STYLE_REV = {1: 'thin', 2: 'medium', 3: 'thick', 4: 'dashed',
                     5: 'dotted', 6: 'double', 7: 'dashDot', 8: 'dashDotDot'}

        def _make_side(b):
            if not b:
                return Side()
            try:
                style_num = int(b.get('style', 1))
            except (TypeError, ValueError):
                style_num = 1
            return Side(border_style=STYLE_REV.get(style_num, 'thin'),
                        color='FF' + str(b.get('color', '#000000')).lstrip('#').upper())

        for bi in (sheet.get('borderInfo') or []):
            if bi.get('rangeType') != 'cell':
                continue
            v = bi.get('value') or {}
            try:
                cell = ws.cell(row=int(v['row_index']) + 1, column=int(v['col_index']) + 1)
            except (KeyError, TypeError, ValueError):
                continue
            cell.border = Border(top=_make_side(v.get('t')), bottom=_make_side(v.get('b')),
                                 left=_make_side(v.get('l')), right=_make_side(v.get('r')))

        for key, m in (sheet.get('merges') or {}).items():
            if m['rs'] > 1 or m['cs'] > 1:
                try:
                    ws.merge_cells(start_row=m['r']+1, start_column=m['c']+1,
                                   end_row=m['r']+m['rs'], end_column=m['c']+m['cs'])
                except Exception:
                    pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='export.xlsx')


@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=False)
